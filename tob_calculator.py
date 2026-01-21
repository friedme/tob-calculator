"""
TOB Calculator Core Module
Handles PDF extraction, ECB rate fetching, and TOB calculations
"""

import xml.etree.ElementTree as ET
import requests
from datetime import datetime, timedelta
from collections import defaultdict
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import csv

# Try importing PDF libraries
try:
    import pdfplumber
    PDF_LIBRARY = 'pdfplumber'
except ImportError:
    try:
        import PyPDF2
        PDF_LIBRARY = 'pypdf2'
    except ImportError:
        PDF_LIBRARY = None

def fetch_ecb_rates(dates_needed):
    """
    Fetch ECB exchange rates from XML feed
    
    Args:
        dates_needed: Set of date strings in 'YYYY-MM-DD' format
    
    Returns:
        dict: {date: {currency: rate}}
    """
    url = "https://www.ecb.europa.eu/stats/eurofxref/eurofxref-hist.xml"
    
    try:
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        xml_content = response.content
    except Exception as e:
        raise Exception(f"Failed to fetch ECB rates: {str(e)}")
    
    # Parse XML
    ns = {
        'gesmes': 'http://www.gesmes.org/xml/2002-08-01',
        'default': 'http://www.ecb.int/vocabulary/2002-08-01/eurofxref'
    }
    
    root = ET.fromstring(xml_content)
    rates = {}
    
    for date_cube in root.findall('.//default:Cube[@time]', ns):
        date = date_cube.get('time')
        if date in dates_needed:
            rates[date] = {'EUR': 1.0}  # EUR to EUR
            for curr_cube in date_cube.findall('default:Cube[@currency]', ns):
                currency = curr_cube.get('currency')
                rate = float(curr_cube.get('rate'))
                rates[date][currency] = rate
    
    # Handle missing dates (weekends/holidays) - use previous business day
    for date in dates_needed:
        if date not in rates:
            rates[date] = get_rate_with_fallback(root, date, ns)
    
    return rates

def get_rate_with_fallback(root, date, ns):
    """Get rate for date, falling back to previous business day if needed"""
    current_date = datetime.strptime(date, '%Y-%m-%d')
    
    # Try up to 5 days back
    for i in range(1, 6):
        prev_date = (current_date - timedelta(days=i)).strftime('%Y-%m-%d')
        
        for date_cube in root.findall('.//default:Cube[@time]', ns):
            if date_cube.get('time') == prev_date:
                rates = {'EUR': 1.0}
                for curr_cube in date_cube.findall('default:Cube[@currency]', ns):
                    currency = curr_cube.get('currency')
                    rate = float(curr_cube.get('rate'))
                    rates[currency] = rate
                return rates
    
    raise ValueError(f"No ECB rate found for {date} or 5 days prior")

def extract_text_from_pdf(pdf_path):
    """Extract text from PDF using available library"""
    if PDF_LIBRARY == 'pdfplumber':
        return extract_with_pdfplumber(pdf_path)
    elif PDF_LIBRARY == 'pypdf2':
        return extract_with_pypdf2(pdf_path)
    else:
        raise Exception("No PDF library available. Install pdfplumber or PyPDF2")

def extract_with_pdfplumber(pdf_path):
    """Extract text using pdfplumber"""
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text += page.extract_text() + "\n"
    return text

def extract_with_pypdf2(pdf_path):
    """Extract text using PyPDF2"""
    text = ""
    with open(pdf_path, 'rb') as file:
        pdf_reader = PyPDF2.PdfReader(file)
        for page in pdf_reader.pages:
            text += page.extract_text() + "\n"
    return text

def detect_broker(text):
    """Detect broker type from PDF text"""
    if 'Interactive Brokers' in text:
        return 'Interactive Brokers'
    elif 'Saxo Bank' in text or 'Transacties' in text:
        return 'Saxo Bank'
    else:
        return 'Unknown'

def extract_ib_transactions(text):
    """
    Extract transactions from Interactive Brokers statement
    Handles the actual IBKR format with currency sections
    """
    transactions = []
    
    lines = text.split('\n')
    in_stocks_section = False
    current_currency = None
    
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        
        # Detect currency sections (JPY, USD, GBP, etc.)
        if line in ['JPY', 'USD', 'GBP', 'EUR', 'CAD', 'AUD', 'SEK', 'NOK', 'CHF', 'HKD', 'SGD']:
            current_currency = line
            in_stocks_section = True
            i += 1
            continue
        
        # Exit stocks section
        if 'Total in GBP' in line or 'Forex' in line or 'Symbol Date/Time Quantity T. Price Proceeds' in line:
            in_stocks_section = False
            current_currency = None
        
        # Parse transaction lines
        if in_stocks_section and current_currency:
            # Look for pattern: YYYY-MM-DD, on one line, then symbol and data on next
            if re.match(r'\d{4}-\d{2}-\d{2},', line):
                date = line.split(',')[0]
                
                # Next line should have: SYMBOL QUANTITY PRICE PRICE PROCEEDS ...
                if i + 1 < len(lines):
                    next_line = lines[i + 1].strip()
                    
                    # Parse: 3836.T -5,000 1,736.0000 1,730.0000 8,680,000 -4,577.06 ...
                    # Or: 4374.T -2,600 2,790.0000 2,825.0000 7,254,000 -4,086.7 ...
                    parts = next_line.split()
                    
                    if len(parts) >= 5:
                        symbol = parts[0]

                        try:
                            quantity_str = parts[1].replace(',', '')
                            quantity = int(quantity_str)

                            # Determine if C.Price column exists by checking if parts[3] looks like a price
                            # Price patterns: X.XXXX (has 4 decimal places)
                            # Proceeds patterns: X,XXX,XXX (large integers, may not have decimals)

                            has_c_price = False
                            if len(parts) >= 4 and '.' in parts[3]:
                                decimal_places = len(parts[3].split('.')[-1])
                                # If 4 decimal places, it's likely C.Price
                                if decimal_places == 4:
                                    has_c_price = True

                            if has_c_price and len(parts) >= 5:
                                # Normal format: SYMBOL QUANTITY T.PRICE C.PRICE PROCEEDS
                                proceeds_str = parts[4].replace(',', '')
                            elif len(parts) >= 4:
                                # Missing C.Price: SYMBOL QUANTITY T.PRICE PROCEEDS
                                proceeds_str = parts[3].replace(',', '')
                            else:
                                continue

                            proceeds = float(proceeds_str)
                            
                            # Skip if this is a "Total" line
                            if symbol.startswith('Total'):
                                i += 1
                                continue
                            
                            # Skip Forex transactions (e.g., USD.JPY, EUR.GBP, etc.)
                            if '.' in symbol and len(symbol.split('.')) == 2:
                                currency_pair = symbol.split('.')
                                # If both parts are 3-letter currency codes, skip
                                if len(currency_pair[0]) == 3 and len(currency_pair[1]) == 3:
                                    i += 1
                                    continue
                            
                            transactions.append({
                                'date': date,
                                'broker': 'Interactive Brokers',
                                'stock': symbol,
                                'type': 'Sell' if quantity < 0 else 'Buy',
                                'shares': abs(quantity),
                                'currency': current_currency,
                                'amount': abs(proceeds)
                            })
                        except (ValueError, IndexError):
                            pass
        
        i += 1
    
    return transactions

def extract_saxo_transactions(text):
    """
    Extract transactions from Saxo Bank "Transactie- en saldorapport" statement
    Parses table format with columns: Transactiedatum, Product, Instrument, Type, Aantal, Koers, Boekingsbedrag
    """
    transactions = []

    lines = text.split('\n')

    for line in lines:
        line_stripped = line.strip()

        # Skip non-stock transactions
        if 'Cashbedrag' in line_stripped or 'Storting/opname' in line_stripped:
            continue

        # Look for transaction lines with date pattern AND 'Aandelen' AND buy/sell
        # Format: 28-nov-2025 01-dec-2025 6494810500 Aandelen JDC Group AG EUR Verkoop SLUITEN -889 26,000 1,0000 ...
        date_match = re.search(r'^(\d{1,2})-(jan|feb|mrt|apr|mei|jun|jul|aug|sep|okt|nov|dec)-(\d{4})', line_stripped, re.IGNORECASE)

        if date_match and 'Aandelen' in line_stripped and ('Koop' in line_stripped or 'Verkoop' in line_stripped):
            day, month_nl, year = date_match.groups()
            month_map = {'jan': 1, 'feb': 2, 'mrt': 3, 'apr': 4, 'mei': 5, 'jun': 6,
                        'jul': 7, 'aug': 8, 'sep': 9, 'okt': 10, 'nov': 11, 'dec': 12}
            month = month_map.get(month_nl.lower(), 1)
            transaction_date = f"{year}-{month:02d}-{int(day):02d}"

            try:
                # Extract instrument name (between Aandelen and currency code)
                # Find "Aandelen" position
                aandelen_idx = line_stripped.find('Aandelen')
                after_aandelen = line_stripped[aandelen_idx + len('Aandelen'):].strip()

                # Instrument name is before the currency code (EUR/USD)
                # Look for EUR, USD, GBP, etc.
                currency_match = re.search(r'\b(EUR|USD|GBP|CAD|AUD|JPY|CHF|SEK|NOK)\b', after_aandelen)
                if not currency_match:
                    continue

                currency = currency_match.group(1)
                instrument_end_idx = currency_match.start()
                instrument = after_aandelen[:instrument_end_idx].strip()

                # Determine transaction type
                trans_type = 'Buy' if 'Koop' in line_stripped else 'Sell'

                # Extract shares (negative number after Koop/Verkoop)
                # Format: "Koop OPENING 655" or "Verkoop SLUITEN -889"
                shares_match = re.search(r'(Koop|Verkoop)\s+\w+\s+(-?\d+(?:\.\d+)?)', line_stripped)
                if not shares_match:
                    # Try alternate format: just number after type
                    shares_match = re.search(r'(Koop|Verkoop).*?(-?\d+(?:\.\d+)?)\s+[\d,]+', line_stripped)

                if not shares_match:
                    continue

                shares = abs(int(float(shares_match.group(2).replace(',', '').replace('.', ''))))

                # Extract amount (Boekingsbedrag column)
                # Look for Belgian format numbers: -26.625,96 or 23.102,44
                # The boekingsbedrag is the largest absolute value (excluding the FX rate which is typically 1,0000)
                amount_matches = re.findall(r'(-?[\d.]+,\d{2})', line_stripped)
                if not amount_matches:
                    continue

                # Find the largest amount (by absolute value), excluding small rates like 1,0000
                boekingsbedrag = None
                max_amount = 0
                for amt_str in amount_matches:
                    # Convert to float
                    amount_str = amt_str.replace('.', '').replace(',', '.').replace('-', '')
                    try:
                        amount_val = float(amount_str)
                        # Skip small values like FX rates (typically around 1.0)
                        if amount_val < 2.0:
                            continue
                        # Keep track of the largest amount
                        if amount_val > max_amount:
                            max_amount = amount_val
                            boekingsbedrag = amount_val
                    except ValueError:
                        continue

                if boekingsbedrag is None or boekingsbedrag == 0:
                    continue

                transactions.append({
                    'date': transaction_date,
                    'broker': 'Saxo Bank',
                    'stock': instrument,
                    'type': trans_type,
                    'shares': shares,
                    'currency': currency,
                    'amount': boekingsbedrag
                })

            except (ValueError, IndexError, AttributeError) as e:
                # Skip lines that don't parse correctly
                continue
    
    return transactions

def group_transactions(transactions):
    """
    Group transactions by date + stock + type
    Same-side transactions (multiple buys or multiple sells) should be grouped
    Opposite-side transactions (day trades) should remain separate
    """
    grouped = defaultdict(lambda: {
        'shares': 0,
        'amount': 0.0,
        'transactions': []
    })
    
    for t in transactions:
        key = (t['date'], t['broker'], t['stock'], t['type'], t['currency'])
        grouped[key]['shares'] += t['shares']
        grouped[key]['amount'] += t['amount']
        grouped[key]['transactions'].append(t)
    
    result = []
    for key, data in grouped.items():
        date, broker, stock, trans_type, currency = key
        result.append({
            'date': date,
            'broker': broker,
            'stock': stock,
            'type': trans_type,
            'shares': data['shares'],
            'currency': currency,
            'amount': data['amount'],
            'grouped_count': len(data['transactions'])
        })
    
    return sorted(result, key=lambda x: (x['date'], x['broker'], x['stock']))

def calculate_tob(transactions, ecb_rates):
    """Calculate TOB for all transactions"""
    results = []
    
    for t in transactions:
        date = t['date']
        currency = t['currency']
        amount = t['amount']
        
        # Get ECB rate
        if date not in ecb_rates or currency not in ecb_rates[date]:
            raise ValueError(f"No ECB rate for {currency} on {date}")
        
        rate = ecb_rates[date][currency]
        
        # Convert to EUR (for EUR, rate is 1.0, so eur_amount = amount)
        eur_amount = round(amount / rate, 2)
        
        # Calculate TOB at 0.35%
        tob = round(eur_amount * 0.0035, 2)
        
        results.append({
            **t,
            'rate': rate,
            'eur_amount': eur_amount,
            'tob': tob
        })
    
    return results

def process_statements(pdf_paths):
    """Main processing function"""
    all_transactions = []
    
    # Extract transactions from all PDFs
    for pdf_path in pdf_paths:
        text = extract_text_from_pdf(pdf_path)
        broker = detect_broker(text)
        
        if broker == 'Interactive Brokers':
            transactions = extract_ib_transactions(text)
        elif broker == 'Saxo Bank':
            transactions = extract_saxo_transactions(text)
        else:
            raise Exception(f"Unknown broker format in {pdf_path}")
        
        all_transactions.extend(transactions)
    
    # Group transactions
    grouped = group_transactions(all_transactions)
    
    # Get unique dates for ECB rate fetching
    dates_needed = set(t['date'] for t in grouped)
    
    # Fetch ECB rates
    ecb_rates = fetch_ecb_rates(dates_needed)
    
    # Calculate TOB
    results = calculate_tob(grouped, ecb_rates)
    
    # Calculate totals
    total_eur = sum(r['eur_amount'] for r in results)
    total_tob = sum(r['tob'] for r in results)
    
    return {
        'transactions': results,
        'total_eur': total_eur,
        'total_tob': total_tob,
        'ecb_rates': ecb_rates
    }

def generate_excel(results, output_path):
    """Generate formatted Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "TOB Tax Report"
    
    # Headers
    headers = ['Date', 'Broker', 'Stock', 'Type', 'Shares', 'Currency', 
               'Amount', 'ECB Rate', 'EUR Amount', 'TOB (0.35%)']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data
    for t in results['transactions']:
        ws.append([
            t['date'],
            t['broker'],
            t['stock'],
            t['type'],
            t['shares'],
            t['currency'],
            t['amount'],
            t['rate'],
            t['eur_amount'],
            t['tob']
        ])
    
    # Add totals row
    last_row = ws.max_row + 1
    ws[f'A{last_row}'] = 'TOTAL'
    ws[f'I{last_row}'] = results['total_eur']
    ws[f'J{last_row}'] = results['total_tob']
    
    # Style totals
    total_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF")
    for col in ['A', 'I', 'J']:
        ws[f'{col}{last_row}'].fill = total_fill
        ws[f'{col}{last_row}'].font = total_font
    
    # Set column widths
    widths = {'A': 12, 'B': 20, 'C': 15, 'D': 8, 'E': 12, 
              'F': 10, 'G': 15, 'H': 12, 'I': 15, 'J': 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    
    # Number formats
    for row in range(2, ws.max_row + 1):
        ws[f'E{row}'].number_format = '#,##0'
        ws[f'G{row}'].number_format = '#,##0.00'
        ws[f'H{row}'].number_format = '0.0000'
        ws[f'I{row}'].number_format = '#,##0.00'
        ws[f'J{row}'].number_format = '#,##0.00'
    
    wb.save(output_path)

def generate_csv(results, output_path):
    """Generate Belgian-formatted CSV"""
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f, delimiter=';')
        
        # Headers
        writer.writerow(['Datum', 'Broker', 'Aandeel', 'Type', 'Aantal', 
                        'Munt', 'Bedrag', 'Koers', 'EUR Bedrag', 'TOB'])
        
        # Data with Belgian number formatting
        for t in results['transactions']:
            writer.writerow([
                t['date'],
                t['broker'],
                t['stock'],
                t['type'],
                format_belgian_number(t['shares']),
                t['currency'],
                format_belgian_number(t['amount']),
                str(t['rate']).replace('.', ','),
                format_belgian_number(t['eur_amount']),
                format_belgian_number(t['tob'])
            ])
        
        # Total row
        writer.writerow([
            'TOTAAL', '', '', '', '', '', '',
            format_belgian_number(results['total_eur']),
            format_belgian_number(results['total_tob'])
        ])

def format_belgian_number(num):
    """Format number in Belgian style: 1.234,56"""
    int_part = int(num)
    dec_part = num - int_part
    
    # Format integer with periods
    int_str = f"{int_part:,}".replace(',', '.')
    
    # Format decimal with comma
    dec_str = f"{dec_part:.2f}"[1:]
    dec_str = dec_str.replace('.', ',')
    
    return int_str + dec_str

def format_belgian_number(num, decimals=2):
    """Format number with Belgian locale: comma as decimal separator, dot as thousands separator"""
    formatted = f"{num:,.{decimals}f}"
    # Replace English formatting with Belgian: swap . and ,
    formatted = formatted.replace(',', 'TEMP').replace('.', ',').replace('TEMP', '.')
    return formatted

def generate_markdown(results, output_path):
    """Generate Markdown summary"""
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("# Belgian TOB Tax Summary\n\n")
        f.write(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write(f"**Total Transactions:** {len(results['transactions'])}\n")
        f.write(f"**Total EUR Amount:** €{results['total_eur']:,.2f}\n")
        f.write(f"**Total TOB Tax:** €{results['total_tob']:,.2f}\n\n")

        f.write("## Transactions\n\n")
        f.write("| Date | Broker | Stock | Type | Shares | Currency | Amount | Rate | EUR Amount | TOB |\n")
        f.write("|------|--------|-------|------|--------|----------|--------|------|------------|-----|\n")

        for t in results['transactions']:
            f.write(f"| {t['date']} | {t['broker']} | {t['stock']} | {t['type']} | "
                   f"{t['shares']:,} | {t['currency']} | {t['amount']:,.2f} | {t['rate']:.4f} | "
                   f"€{t['eur_amount']:,.2f} | €{t['tob']:.2f} |\n")

        f.write("\n## Methodology\n\n")
        f.write("- TOB Rate: 0.35%\n")
        f.write("- Exchange rates: ECB official rates\n")
        f.write("- Grouping: Same stock + same date + same type = 1 transaction\n")
        f.write("- Day trades: Buy and sell transactions kept separate (both taxed)\n")

def generate_pdf(results, output_path):
    """Generate PDF summary report"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                           leftMargin=2*cm, rightMargin=2*cm,
                           topMargin=2*cm, bottomMargin=2*cm)

    story = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#366092'),
        alignment=TA_CENTER,
        spaceAfter=20
    )
    story.append(Paragraph("Belgian TOB Tax Report", title_style))
    story.append(Spacer(1, 0.5*cm))

    # Summary info
    summary_style = ParagraphStyle(
        'Summary',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6
    )
    story.append(Paragraph(f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", summary_style))
    story.append(Paragraph(f"<b>Total Transactions:</b> {len(results['transactions'])}", summary_style))
    story.append(Paragraph(f"<b>Total EUR Amount:</b> €{format_belgian_number(results['total_eur'])}", summary_style))
    story.append(Paragraph(f"<b>Total TOB Tax (0.35%):</b> €{format_belgian_number(results['total_tob'])}", summary_style))
    story.append(Spacer(1, 1*cm))

    # Transactions table
    story.append(Paragraph("<b>Transaction Details</b>", styles['Heading2']))
    story.append(Spacer(1, 0.3*cm))

    # Table data
    table_data = [['Date', 'Broker', 'Stock', 'Type', 'Shares', 'Currency',
                   'Amount', 'Rate', 'EUR Amount', 'TOB']]

    for t in results['transactions']:
        table_data.append([
            t['date'],
            t['broker'],
            t['stock'],
            t['type'],
            format_belgian_number(t['shares'], 0),
            t['currency'],
            format_belgian_number(t['amount']),
            format_belgian_number(t['rate'], 4),
            f"€{format_belgian_number(t['eur_amount'])}",
            f"€{format_belgian_number(t['tob'])}"
        ])

    # Add total row
    table_data.append([
        '', '', '', '', '', '', '', 'TOTAL:',
        f"€{format_belgian_number(results['total_eur'])}",
        f"€{format_belgian_number(results['total_tob'])}"
    ])

    # Create table
    table = Table(table_data, colWidths=[2*cm, 3*cm, 2*cm, 1.5*cm, 1.5*cm,
                                         1.5*cm, 2*cm, 1.5*cm, 2*cm, 1.8*cm])

    # Style table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f0f0f0')])
    ]))

    story.append(table)
    story.append(Spacer(1, 1*cm))

    # Methodology
    story.append(Paragraph("<b>Methodology</b>", styles['Heading2']))
    story.append(Spacer(1, 0.3*cm))
    methodology_text = """
    <b>TOB Rate:</b> 0.35% (maximum €1,600 per transaction)<br/>
    <b>Exchange Rates:</b> Official ECB (European Central Bank) rates<br/>
    <b>Transaction Grouping:</b> Same stock + same date + same type = 1 transaction<br/>
    <b>Buy and Sell on Same Day:</b> Both transactions are kept separate and taxed individually<br/>
    """
    story.append(Paragraph(methodology_text, styles['Normal']))

    story.append(Spacer(1, 0.5*cm))
    disclaimer_style = ParagraphStyle(
        'Disclaimer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    story.append(Paragraph(
        "<i>This report is for informational purposes only. Always verify calculations with a qualified tax professional.</i>",
        disclaimer_style
    ))

    # Build PDF
    doc.build(story)
