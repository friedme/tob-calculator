"""
TOB Output Generators
Creates Excel, CSV, PDF, and Markdown output files for TOB tax reports
All outputs use Belgian number formatting (period for thousands, comma for decimals)
"""

import csv
from datetime import datetime
from typing import List, Dict, Any
from pathlib import Path

# Excel library
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

# PDF library
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm, cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


# =============================================================================
# BELGIAN NUMBER FORMATTING
# =============================================================================

def format_belgian(num: float, decimals: int = 2) -> str:
    """
    Format number in Belgian style: 1.234,56
    
    Args:
        num: Number to format
        decimals: Number of decimal places (default 2)
        
    Returns:
        Belgian-formatted string
    """
    if num is None:
        return ""
    
    # Handle negative numbers
    sign = "-" if num < 0 else ""
    num = abs(num)
    
    if decimals == 0:
        int_part = int(round(num))
        formatted = f"{int_part:,}".replace(',', '.')
    else:
        int_part = int(num)
        dec_part = round(num - int_part, decimals)
        
        # Format integer with periods as thousands separator
        int_str = f"{int_part:,}".replace(',', '.')
        
        # Format decimal with comma
        dec_str = f"{dec_part:.{decimals}f}"[1:]  # Get ".XX" part
        dec_str = dec_str.replace('.', ',')  # Change to ",XX"
        
        formatted = int_str + dec_str
    
    return sign + formatted


def format_belgian_rate(rate: float) -> str:
    """Format ECB rate with 4 decimals, Belgian style"""
    return str(round(rate, 4)).replace('.', ',')


def format_eur(amount: float) -> str:
    """Format as EUR amount: € 1.234,56"""
    return f"€ {format_belgian(amount)}"


# =============================================================================
# EXCEL OUTPUT
# =============================================================================

def generate_excel(results: Dict[str, Any], output_path: str) -> str:
    """
    Generate formatted Excel file with TOB calculations
    
    Args:
        results: Dictionary with transactions, totals, and ECB rates
        output_path: Path for output file
        
    Returns:
        Path to created file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "TOB Berekening"
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    
    subheader_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    
    total_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    total_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    
    data_font = Font(size=10, name="Arial")
    number_font = Font(size=10, name="Arial")
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Title row
    ws.merge_cells('A1:J1')
    ws['A1'] = "Belgische Taks op Beursverrichtingen (TOB) - Berekening"
    ws['A1'].font = Font(bold=True, size=14, name="Arial")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Info row
    ws.merge_cells('A2:J2')
    ws['A2'] = f"Gegenereerd op: {datetime.now().strftime('%d-%m-%Y %H:%M')} | Periode: November 2025"
    ws['A2'].font = Font(size=10, italic=True, name="Arial", color="666666")
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # Empty row
    current_row = 4
    
    # Headers
    headers = [
        ('Datum', 12),
        ('Broker', 18),
        ('Aandeel', 15),
        ('Type', 8),
        ('Aantal', 12),
        ('Munt', 8),
        ('Bedrag', 15),
        ('ECB Koers', 12),
        ('EUR Bedrag', 15),
        ('TOB (0,35%)', 12)
    ]
    
    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = width
    
    ws.row_dimensions[current_row].height = 25
    current_row += 1
    
    # Data rows
    transactions = results.get('transactions', [])
    
    # Group by broker for subtotals
    brokers = {}
    for t in transactions:
        broker = t.get('broker', 'Unknown')
        if broker not in brokers:
            brokers[broker] = []
        brokers[broker].append(t)
    
    for broker, broker_transactions in brokers.items():
        # Broker subheader
        ws.merge_cells(f'A{current_row}:J{current_row}')
        ws[f'A{current_row}'] = broker
        ws[f'A{current_row}'].fill = subheader_fill
        ws[f'A{current_row}'].font = subheader_font
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1
        
        broker_eur_total = 0
        broker_tob_total = 0
        
        for t in broker_transactions:
            # Date
            ws.cell(row=current_row, column=1, value=t.get('date', ''))
            ws.cell(row=current_row, column=1).font = data_font
            ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center')
            
            # Broker
            ws.cell(row=current_row, column=2, value=t.get('broker', ''))
            ws.cell(row=current_row, column=2).font = data_font
            
            # Stock
            ws.cell(row=current_row, column=3, value=t.get('stock', ''))
            ws.cell(row=current_row, column=3).font = data_font
            
            # Type
            trans_type = t.get('type', '')
            cell = ws.cell(row=current_row, column=4, value=trans_type)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center')
            if trans_type == 'Buy':
                cell.font = Font(size=10, name="Arial", color="006600")
            else:
                cell.font = Font(size=10, name="Arial", color="CC0000")
            
            # Shares
            shares = t.get('shares', 0)
            ws.cell(row=current_row, column=5, value=shares)
            ws.cell(row=current_row, column=5).font = number_font
            ws.cell(row=current_row, column=5).number_format = '#,##0'
            ws.cell(row=current_row, column=5).alignment = Alignment(horizontal='right')
            
            # Currency
            ws.cell(row=current_row, column=6, value=t.get('currency', ''))
            ws.cell(row=current_row, column=6).font = data_font
            ws.cell(row=current_row, column=6).alignment = Alignment(horizontal='center')
            
            # Amount (original currency)
            amount = t.get('amount', 0)
            ws.cell(row=current_row, column=7, value=amount)
            ws.cell(row=current_row, column=7).font = number_font
            ws.cell(row=current_row, column=7).number_format = '#,##0.00'
            ws.cell(row=current_row, column=7).alignment = Alignment(horizontal='right')
            
            # ECB Rate
            rate = t.get('rate', 1.0)
            ws.cell(row=current_row, column=8, value=rate)
            ws.cell(row=current_row, column=8).font = number_font
            ws.cell(row=current_row, column=8).number_format = '0.0000'
            ws.cell(row=current_row, column=8).alignment = Alignment(horizontal='right')
            
            # EUR Amount
            eur_amount = t.get('eur_amount', 0)
            ws.cell(row=current_row, column=9, value=eur_amount)
            ws.cell(row=current_row, column=9).font = number_font
            ws.cell(row=current_row, column=9).number_format = '€ #,##0.00'
            ws.cell(row=current_row, column=9).alignment = Alignment(horizontal='right')
            broker_eur_total += eur_amount
            
            # TOB
            tob = t.get('tob', 0)
            ws.cell(row=current_row, column=10, value=tob)
            ws.cell(row=current_row, column=10).font = number_font
            ws.cell(row=current_row, column=10).number_format = '€ #,##0.00'
            ws.cell(row=current_row, column=10).alignment = Alignment(horizontal='right')
            broker_tob_total += tob
            
            # Apply borders
            for col in range(1, 11):
                ws.cell(row=current_row, column=col).border = thin_border
            
            current_row += 1
        
        # Broker subtotal
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = f"Subtotaal {broker}"
        ws[f'A{current_row}'].font = Font(bold=True, size=10, name="Arial")
        ws[f'A{current_row}'].alignment = Alignment(horizontal='right')
        
        ws.cell(row=current_row, column=9, value=broker_eur_total)
        ws.cell(row=current_row, column=9).font = Font(bold=True, size=10, name="Arial")
        ws.cell(row=current_row, column=9).number_format = '€ #,##0.00'
        ws.cell(row=current_row, column=9).alignment = Alignment(horizontal='right')
        
        ws.cell(row=current_row, column=10, value=broker_tob_total)
        ws.cell(row=current_row, column=10).font = Font(bold=True, size=10, name="Arial")
        ws.cell(row=current_row, column=10).number_format = '€ #,##0.00'
        ws.cell(row=current_row, column=10).alignment = Alignment(horizontal='right')
        
        for col in range(1, 11):
            ws.cell(row=current_row, column=col).border = thin_border
            ws.cell(row=current_row, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        current_row += 2
    
    # Grand total row
    ws.merge_cells(f'A{current_row}:H{current_row}')
    ws[f'A{current_row}'] = "TOTAAL TOB VERSCHULDIGD"
    ws[f'A{current_row}'].fill = total_fill
    ws[f'A{current_row}'].font = total_font
    ws[f'A{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
    
    ws.cell(row=current_row, column=9, value=results.get('total_eur', 0))
    ws.cell(row=current_row, column=9).fill = total_fill
    ws.cell(row=current_row, column=9).font = total_font
    ws.cell(row=current_row, column=9).number_format = '€ #,##0.00'
    ws.cell(row=current_row, column=9).alignment = Alignment(horizontal='right')
    
    ws.cell(row=current_row, column=10, value=results.get('total_tob', 0))
    ws.cell(row=current_row, column=10).fill = total_fill
    ws.cell(row=current_row, column=10).font = total_font
    ws.cell(row=current_row, column=10).number_format = '€ #,##0.00'
    ws.cell(row=current_row, column=10).alignment = Alignment(horizontal='right')
    
    ws.row_dimensions[current_row].height = 25
    
    for col in range(1, 11):
        ws.cell(row=current_row, column=col).border = thin_border
    
    # Summary section
    current_row += 3
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "Samenvatting"
    ws[f'A{current_row}'].font = Font(bold=True, size=12, name="Arial")
    
    current_row += 1
    summary_data = [
        ("Aantal transacties:", len(transactions)),
        ("Totaal EUR bedrag:", format_eur(results.get('total_eur', 0))),
        ("TOB tarief:", "0,35%"),
        ("Totaal TOB:", format_eur(results.get('total_tob', 0))),
    ]
    
    for label, value in summary_data:
        ws.cell(row=current_row, column=1, value=label)
        ws.cell(row=current_row, column=1).font = Font(size=10, name="Arial")
        ws.cell(row=current_row, column=2, value=value if isinstance(value, str) else value)
        ws.cell(row=current_row, column=2).font = Font(bold=True, size=10, name="Arial")
        current_row += 1
    
    
    # Freeze panes
    ws.freeze_panes = 'A5'
    
    # Save
    wb.save(output_path)
    return output_path


# =============================================================================
# CSV OUTPUT (BELGIAN FORMAT)
# =============================================================================

def generate_csv(results: Dict[str, Any], output_path: str) -> str:
    """
    Generate Belgian-formatted CSV file
    
    Belgian CSV format:
    - Semicolon (;) as delimiter
    - Comma (,) as decimal separator
    - Period (.) as thousands separator
    
    Args:
        results: Dictionary with transactions and totals
        output_path: Path for output file
        
    Returns:
        Path to created file
    """
    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f, delimiter=';')
        
        # Headers (Dutch)
        writer.writerow([
            'Datum', 'Broker', 'Aandeel', 'Type', 'Aantal',
            'Munt', 'Bedrag', 'ECB Koers', 'EUR Bedrag', 'TOB'
        ])
        
        # Data rows with Belgian formatting
        for t in results.get('transactions', []):
            writer.writerow([
                t.get('date', ''),
                t.get('broker', ''),
                t.get('stock', ''),
                t.get('type', ''),
                format_belgian(t.get('shares', 0), 0),
                t.get('currency', ''),
                format_belgian(t.get('amount', 0)),
                format_belgian_rate(t.get('rate', 1.0)),
                format_belgian(t.get('eur_amount', 0)),
                format_belgian(t.get('tob', 0))
            ])
        
        # Empty row
        writer.writerow([])
        
        # Totals
        writer.writerow([
            'TOTAAL', '', '', '', '', '', '',
            '',
            format_belgian(results.get('total_eur', 0)),
            format_belgian(results.get('total_tob', 0))
        ])
        
        # Summary info
        writer.writerow([])
        writer.writerow(['Samenvatting'])
        writer.writerow(['Aantal transacties', len(results.get('transactions', []))])
        writer.writerow(['Totaal EUR bedrag', format_belgian(results.get('total_eur', 0))])
        writer.writerow(['TOB tarief', '0,35%'])
        writer.writerow(['Totaal TOB', format_belgian(results.get('total_tob', 0))])
    
    return output_path


# =============================================================================
# PDF OUTPUT
# =============================================================================

def generate_pdf(results: Dict[str, Any], output_path: str) -> str:
    """
    Generate PDF report with TOB calculations
    
    Args:
        results: Dictionary with transactions and totals
        output_path: Path for output file
        
    Returns:
        Path to created file
        
    Raises:
        ImportError: If reportlab is not available
    """
    if not REPORTLAB_AVAILABLE:
        raise ImportError(
            "reportlab is required for PDF generation. "
            "Install with: pip install reportlab"
        )
    
    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    # Styles
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=20,
        alignment=1  # Center
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=20,
        alignment=1
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=12,
        spaceBefore=20,
        spaceAfter=10
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10
    )
    
    # Build story
    story = []
    
    # Title
    story.append(Paragraph(
        "Belgische Taks op Beursverrichtingen (TOB)",
        title_style
    ))
    
    story.append(Paragraph(
        f"Gegenereerd op {datetime.now().strftime('%d-%m-%Y %H:%M')}",
        subtitle_style
    ))
    
    # Summary box
    story.append(Paragraph("Samenvatting", heading_style))
    
    summary_data = [
        ["Aantal transacties", str(len(results.get('transactions', [])))],
        ["Totaal EUR bedrag", format_eur(results.get('total_eur', 0))],
        ["TOB tarief", "0,35%"],
        ["Totaal TOB verschuldigd", format_eur(results.get('total_tob', 0))],
    ]
    
    summary_table = Table(summary_data, colWidths=[150, 150])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.Color(0.95, 0.95, 0.95)),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (1, 3), (1, 3), 'Helvetica-Bold'),
        ('FONTSIZE', (1, 3), (1, 3), 12),
        ('TEXTCOLOR', (1, 3), (1, 3), colors.Color(0.12, 0.31, 0.47)),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.8, 0.8, 0.8)),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Transactions table
    story.append(Paragraph("Transacties", heading_style))
    
    # Table headers
    table_headers = [
        'Datum', 'Broker', 'Aandeel', 'Type', 'Aantal',
        'Munt', 'Bedrag', 'Koers', 'EUR', 'TOB'
    ]
    
    table_data = [table_headers]
    
    for t in results.get('transactions', []):
        row = [
            t.get('date', ''),
            t.get('broker', '')[:15],  # Truncate for space
            t.get('stock', '')[:12],
            t.get('type', ''),
            format_belgian(t.get('shares', 0), 0),
            t.get('currency', ''),
            format_belgian(t.get('amount', 0)),
            format_belgian_rate(t.get('rate', 1.0)),
            format_belgian(t.get('eur_amount', 0)),
            format_belgian(t.get('tob', 0))
        ]
        table_data.append(row)
    
    # Total row
    table_data.append([
        'TOTAAL', '', '', '', '', '', '', '',
        format_belgian(results.get('total_eur', 0)),
        format_belgian(results.get('total_tob', 0))
    ])
    
    # Create table with appropriate column widths
    col_widths = [55, 65, 55, 35, 45, 30, 55, 40, 50, 45]
    
    trans_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    trans_table.setStyle(TableStyle([
        # Header style
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.12, 0.31, 0.47)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        # Data style
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 7),
        ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),
        ('ALIGN', (5, 1), (5, -1), 'CENTER'),
        
        # Total row style
        ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.12, 0.31, 0.47)),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 8),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.8, 0.8, 0.8)),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        
        # Alternating row colors
        *[('BACKGROUND', (0, i), (-1, i), colors.Color(0.97, 0.97, 0.97)) 
          for i in range(2, len(table_data)-1, 2)]
    ]))
    story.append(trans_table)
    
    
    
    # Footer note
    story.append(Spacer(1, 30))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey
    )
    story.append(Paragraph(
        "Dit document is automatisch gegenereerd. ECB wisselkoersen zijn afkomstig van "
        "de officiële ECB XML feed. TOB wordt berekend aan 0,35% op het EUR bedrag.",
        footer_style
    ))
    
    # Build PDF
    doc.build(story)
    return output_path


# =============================================================================
# MARKDOWN OUTPUT
# =============================================================================

def generate_markdown(results: Dict[str, Any], output_path: str) -> str:
    """
    Generate Markdown summary file
    
    Args:
        results: Dictionary with transactions and totals
        output_path: Path for output file
        
    Returns:
        Path to created file
    """
    transactions = results.get('transactions', [])
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("# Belgian TOB Tax Summary\n\n")
        f.write("## Belgische Taks op Beursverrichtingen\n\n")
        
        f.write(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        
        # Summary
        f.write("## Summary\n\n")
        f.write(f"| Metric | Value |\n")
        f.write(f"|--------|-------|\n")
        f.write(f"| Number of transactions | {len(transactions)} |\n")
        f.write(f"| Total EUR amount | {format_eur(results.get('total_eur', 0))} |\n")
        f.write(f"| TOB rate | 0,35% |\n")
        f.write(f"| **Total TOB due** | **{format_eur(results.get('total_tob', 0))}** |\n\n")
        
        # Transactions by broker
        f.write("## Transactions\n\n")
        
        # Group by broker
        brokers = {}
        for t in transactions:
            broker = t.get('broker', 'Unknown')
            if broker not in brokers:
                brokers[broker] = []
            brokers[broker].append(t)
        
        for broker, broker_trans in brokers.items():
            f.write(f"### {broker}\n\n")
            
            f.write("| Date | Stock | Type | Shares | Currency | Amount | Rate | EUR Amount | TOB |\n")
            f.write("|------|-------|------|--------|----------|--------|------|------------|-----|\n")
            
            broker_total_eur = 0
            broker_total_tob = 0
            
            for t in broker_trans:
                eur_amount = t.get('eur_amount', 0)
                tob = t.get('tob', 0)
                broker_total_eur += eur_amount
                broker_total_tob += tob
                
                f.write(f"| {t.get('date', '')} ")
                f.write(f"| {t.get('stock', '')} ")
                f.write(f"| {t.get('type', '')} ")
                f.write(f"| {t.get('shares', 0):,} ")
                f.write(f"| {t.get('currency', '')} ")
                f.write(f"| {format_belgian(t.get('amount', 0))} ")
                f.write(f"| {format_belgian_rate(t.get('rate', 1.0))} ")
                f.write(f"| {format_eur(eur_amount)} ")
                f.write(f"| {format_eur(tob)} |\n")
            
            f.write(f"| **Subtotal** | | | | | | | **{format_eur(broker_total_eur)}** | **{format_eur(broker_total_tob)}** |\n\n")
        
        # ECB Rates
        f.write("## ECB Exchange Rates Used\n\n")
        
        ecb_rates = results.get('ecb_rates', {})
        currencies_used = set()
        for t in transactions:
            if t.get('currency') != 'EUR':
                currencies_used.add(t.get('currency'))
        
        if currencies_used:
            f.write("| Date | Currency | Rate |\n")
            f.write("|------|----------|------|\n")
            
            for date in sorted(ecb_rates.keys()):
                rates = ecb_rates[date]
                for curr in sorted(currencies_used):
                    if curr in rates:
                        f.write(f"| {date} | {curr} | {format_belgian_rate(rates[curr])} |\n")
            
            f.write("\n")
        
        # Methodology
        f.write("## Methodology\n\n")
        f.write("- **TOB Rate:** 0,35% (0.0035)\n")
        f.write("- **Exchange rates:** Official ECB rates from XML feed\n")
        f.write("- **Conversion formula:** EUR = Amount ÷ ECB Rate\n")
        f.write("- **Grouping:** Same stock + same date + same type = 1 transaction\n")
        f.write("- **Day trades:** Buy and sell transactions kept separate (both taxed)\n")
        f.write("- **Maximum TOB:** € 1.600 per transaction\n\n")
        
        f.write("---\n")
        f.write(f"*Source: ECB XML feed - https://www.ecb.europa.eu/stats/eurofxref/eurofxref-hist.xml*\n")
    
    return output_path


# =============================================================================
# CONVENIENCE FUNCTION
# =============================================================================

def generate_all_outputs(results: Dict[str, Any], output_dir: str, base_name: str = "tob_report") -> Dict[str, str]:
    """
    Generate all output formats (Excel, CSV, PDF, Markdown)
    
    Args:
        results: Dictionary with transactions and totals
        output_dir: Directory for output files
        base_name: Base name for output files
        
    Returns:
        Dictionary mapping format to file path
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    outputs = {}
    
    # Excel
    excel_path = output_dir / f"{base_name}.xlsx"
    outputs['excel'] = generate_excel(results, str(excel_path))
    
    # CSV
    csv_path = output_dir / f"{base_name}.csv"
    outputs['csv'] = generate_csv(results, str(csv_path))
    
    # PDF (if available)
    if REPORTLAB_AVAILABLE:
        pdf_path = output_dir / f"{base_name}.pdf"
        outputs['pdf'] = generate_pdf(results, str(pdf_path))
    
    # Markdown
    md_path = output_dir / f"{base_name}.md"
    outputs['markdown'] = generate_markdown(results, str(md_path))
    
    return outputs
