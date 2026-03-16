"""
TOB Core Module - PDF Extraction and Processing
Handles extraction of stock transactions from Belgian broker statements
Supports: Interactive Brokers (English) and Saxo Bank (Dutch/Flemish)
"""

import re
import xml.etree.ElementTree as ET
import requests
from datetime import datetime, timedelta
from collections import defaultdict
from dataclasses import dataclass
from typing import List, Dict, Optional, Tuple, Any
from enum import Enum

# Try importing PDF libraries
try:
    import pdfplumber
    PDF_LIBRARY = 'pdfplumber'
except ImportError:
    try:
        from pypdf import PdfReader
        PDF_LIBRARY = 'pypdf'
    except ImportError:
        PDF_LIBRARY = None


class Broker(Enum):
    INTERACTIVE_BROKERS = "Interactive Brokers"
    SAXO_BANK = "Saxo Bank"
    UNKNOWN = "Unknown"


class ExtractionError(Exception):
    """Custom exception for extraction errors with detailed messages"""
    def __init__(self, message: str, error_type: str = "extraction_error", details: dict = None):
        super().__init__(message)
        self.error_type = error_type
        self.details = details or {}


@dataclass
class Transaction:
    """Represents a single stock transaction"""
    date: str              # YYYY-MM-DD
    broker: str
    stock: str
    trans_type: str        # 'Buy' or 'Sell'
    shares: int
    currency: str
    amount: float
    original_line: str = ""  # For debugging


@dataclass
class ExtractionResult:
    """Result of PDF extraction"""
    transactions: List[Transaction]
    broker: Broker
    warnings: List[str]
    raw_text: str
    success: bool
    error_message: str = ""


# =============================================================================
# PDF TEXT EXTRACTION
# =============================================================================

def extract_text_from_pdf(pdf_path: str) -> str:
    """
    Extract text from PDF using available library
    
    Args:
        pdf_path: Path to PDF file
        
    Returns:
        Extracted text content
        
    Raises:
        ExtractionError: If no PDF library is available or extraction fails
    """
    if PDF_LIBRARY is None:
        raise ExtractionError(
            "No PDF library available. Install pdfplumber: pip install pdfplumber",
            "library_missing"
        )
    
    try:
        if PDF_LIBRARY == 'pdfplumber':
            return _extract_with_pdfplumber(pdf_path)
        else:
            return _extract_with_pypdf(pdf_path)
    except Exception as e:
        raise ExtractionError(
            f"Failed to read PDF: {str(e)}",
            "pdf_read_error",
            {"path": pdf_path, "original_error": str(e)}
        )


def _extract_with_pdfplumber(pdf_path: str) -> str:
    """Extract text using pdfplumber (preserves layout better)"""
    import pdfplumber
    text_parts = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
    return "\n".join(text_parts)


def _extract_with_pypdf(pdf_path: str) -> str:
    """Extract text using pypdf"""
    from pypdf import PdfReader
    reader = PdfReader(pdf_path)
    text_parts = []
    for page in reader.pages:
        page_text = page.extract_text()
        if page_text:
            text_parts.append(page_text)
    return "\n".join(text_parts)


# =============================================================================
# BROKER DETECTION
# =============================================================================

def detect_broker(text: str) -> Broker:
    """
    Detect broker type from PDF text content
    
    Args:
        text: Extracted PDF text
        
    Returns:
        Broker enum value
    """
    text_lower = text.lower()
    
    # Interactive Brokers markers
    ib_markers = [
        'interactive brokers',
        'activity statement',
        'trades',
        'realized & unrealized',
        'mark-to-market'
    ]
    
    # Saxo Bank markers (Dutch/Flemish)
    saxo_markers = [
        'saxo bank',
        'transacties',
        'transactie- en saldorapport',
        'zelf beleggen',
        'boekingsbedrag'
    ]
    
    ib_score = sum(1 for marker in ib_markers if marker in text_lower)
    saxo_score = sum(1 for marker in saxo_markers if marker in text_lower)
    
    if ib_score >= 2:
        return Broker.INTERACTIVE_BROKERS
    elif saxo_score >= 2:
        return Broker.SAXO_BANK
    else:
        return Broker.UNKNOWN


# =============================================================================
# INTERACTIVE BROKERS EXTRACTION
# =============================================================================
def extract_ib_transactions(text: str) -> ExtractionResult:
    """
    Extract transactions from Interactive Brokers statement
    
    Format (date on separate line):
    Line 1: 2025-11-26,
    Line 2: MCE -350,000 0.2150 0.2300 75,250.00 -61.44 ...
    Line 3: 18:00:07
    
    CRITICAL: 
    - JPY stocks have INTEGER proceeds (no decimals)
    - Other currencies have DECIMAL proceeds
    - Symbols can be mixed case (e.g., ZEGl)
    """
    transactions = []
    warnings = []
    
    lines = text.split('\n')
    
    current_currency = None
    in_trades_section = False
    current_date = None
    
    # Pattern to detect Trades section start
    trades_section_pattern = re.compile(r'^Trades\s*$', re.IGNORECASE)
    
    # Pattern for date line: 2025-11-26,
    date_pattern = re.compile(r'^(\d{4}-\d{2}-\d{2}),')
    
    # Pattern for JPY stocks - proceeds are INTEGERS (no decimal point)
    # Japanese stocks always end in .T
    jpy_transaction_pattern = re.compile(
        r'^(\d+\.T)\s+'                  # Symbol (JPY stocks end in .T)
        r'(-?[\d,]+)\s+'                 # Quantity
        r'[\d,.]+\s+'                    # T.Price
        r'(?:[\d,.]+\s+)?'               # Optional C.Price
        r'(-?[\d,]+)\s+'                 # Proceeds (INTEGER - no decimal!)
        r'(-?[\d,.]+)'                   # Comm/Fee (has decimal)
    )
    
    # Pattern for non-JPY stocks - proceeds HAVE decimals
    # FIXED: Allow mixed case symbols (e.g., ZEGl)
    regular_transaction_pattern = re.compile(
        r'^([A-Za-z0-9.]+)\s+'           # Symbol - MIXED CASE!
        r'(-?[\d,]+)\s+'                 # Quantity
        r'[\d,.]+\s+'                    # T.Price
        r'(?:[\d,.]+|--)\s+'             # C.Price
        r'(-?[\d,]+\.\d+)\s+'            # Proceeds (HAS decimal point!)
        r'(-?[\d,.]+)'                   # Comm/Fee
    )
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Check for Trades section start
        if trades_section_pattern.match(line):
            in_trades_section = True
            continue
        
        if not in_trades_section:
            continue
        
        # Check for currency headers
        for curr in ['AUD', 'CAD', 'GBP', 'JPY', 'USD', 'EUR', 'SEK', 'CHF', 'HKD', 'NOK', 'DKK']:
            if line == curr or line == f'Stocks{curr}':
                current_currency = curr
                break
        
        # Check for date line
        date_match = date_pattern.match(line)
        if date_match:
            current_date = date_match.group(1)
            continue
        
        # Check for section end markers
        if any(marker in line.lower() for marker in ['forex', 'equity and index options']):
            break
        
        # Skip Total lines
        if line.startswith('Total'):
            continue
        
        # Try JPY pattern first (for Japanese stocks)
        jpy_match = jpy_transaction_pattern.match(line)
        if jpy_match and current_date:
            symbol, quantity_str, proceeds_str, comm_str = jpy_match.groups()
            
            try:
                quantity = int(quantity_str.replace(',', ''))
                proceeds = abs(float(proceeds_str.replace(',', '')))
            except ValueError:
                continue
            
            trans_type = 'Sell' if quantity < 0 else 'Buy'
            
            transactions.append(Transaction(
                date=current_date,
                broker=Broker.INTERACTIVE_BROKERS.value,
                stock=symbol,
                trans_type=trans_type,
                shares=abs(quantity),
                currency='JPY',  # Japanese stocks are always JPY
                amount=proceeds,
                original_line=line
            ))
            continue
        
        # Try regular pattern for non-JPY stocks
        regular_match = regular_transaction_pattern.match(line)
        if regular_match and current_date:
            symbol, quantity_str, proceeds_str, comm_str = regular_match.groups()
            
            # Skip Total lines
            if 'Total' in symbol:
                continue
            
            try:
                quantity = int(quantity_str.replace(',', ''))
                proceeds = abs(float(proceeds_str.replace(',', '')))
            except ValueError:
                continue
            
            trans_type = 'Sell' if quantity < 0 else 'Buy'
            
            # Determine currency
            trade_currency = current_currency
            if not trade_currency:
                # Look back for currency context
                context = ' '.join(lines[max(0, i-10):i])
                for curr in ['AUD', 'CAD', 'GBP', 'USD', 'EUR', 'SEK']:
                    if curr in context:
                        trade_currency = curr
                        break
            
            if not trade_currency:
                warnings.append(f"Could not determine currency for {symbol} on {current_date}, defaulting to USD")
                trade_currency = 'USD'
            
            transactions.append(Transaction(
                date=current_date,
                broker=Broker.INTERACTIVE_BROKERS.value,
                stock=symbol,
                trans_type=trans_type,
                shares=abs(quantity),
                currency=trade_currency,
                amount=proceeds,
                original_line=line
            ))
    
    # Alternative extraction if pattern-based didn't work well
    if not transactions:
        transactions, alt_warnings = _extract_ib_alternative(text)
        warnings.extend(alt_warnings)
    
    if not transactions:
        return ExtractionResult(
            transactions=[],
            broker=Broker.INTERACTIVE_BROKERS,
            warnings=warnings,
            raw_text=text,
            success=False,
            error_message="No transactions found in Interactive Brokers statement."
        )
    
    return ExtractionResult(
        transactions=transactions,
        broker=Broker.INTERACTIVE_BROKERS,
        warnings=warnings,
        raw_text=text,
        success=True
    )

def _extract_ib_alternative(text: str) -> Tuple[List[Transaction], List[str]]:
    """
    Alternative extraction method for IB statements
    Uses more flexible pattern matching for different PDF layouts
    """
    transactions = []
    warnings = []
    
    # Look for the Trades table specifically
    # Pattern for trades with date, symbol, and numbers
    # More flexible pattern that captures the key elements
    
    flexible_pattern = re.compile(
        r'([A-Z0-9.]+(?:\.[A-Z]+)?)\s+'   # Symbol
        r'(\d{4}-\d{2}-\d{2})'             # Date
        r'[,\s]+[\d:]+\s+'                 # Time separator
        r'(-?[\d,]+)\s+'                   # Quantity
        r'[\d.]+\s+'                       # Price
        r'[\d.]+\s+'                       # Another price or --
        r'(-?[\d,]+\.?\d*)'                # Amount (proceeds)
    )
    
    # Detect currency sections
    current_currency = 'USD'  # Default
    lines = text.split('\n')
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Update currency context
        for curr in ['AUD', 'CAD', 'GBP', 'JPY', 'USD', 'EUR', 'SEK']:
            if f'Stocks{curr}' in line or line == curr:
                current_currency = curr
                break
        
        # Japanese stocks
        if re.search(r'\d+\.T', line):
            current_currency = 'JPY'
        
        match = flexible_pattern.search(line)
        if match:
            symbol, date, quantity_str, amount_str = match.groups()
            
            # Skip headers and non-trade lines
            if symbol.upper() in ['SYMBOL', 'TOTAL', 'SUBTOTAL']:
                continue
            
            quantity = int(quantity_str.replace(',', ''))
            amount = abs(float(amount_str.replace(',', '')))
            
            trans_type = 'Sell' if quantity < 0 else 'Buy'
            
            # Override currency for Japanese stocks
            trade_currency = current_currency
            if re.match(r'^\d+\.T$', symbol):
                trade_currency = 'JPY'
            
            transactions.append(Transaction(
                date=date,
                broker=Broker.INTERACTIVE_BROKERS.value,
                stock=symbol,
                trans_type=trans_type,
                shares=abs(quantity),
                currency=trade_currency,
                amount=amount,
                original_line=line
            ))
    
    return transactions, warnings


# =============================================================================
# SAXO BANK EXTRACTION
# =============================================================================

def _get_instrument_from_adjacent_lines(lines: list, current_idx: int) -> str:
    """
    Reconstruct instrument name from adjacent lines when pdfplumber splits
    long instrument names across multiple lines.

    Saxo PDFs place the first part of the name on the line BEFORE the transaction
    and the second part on the line AFTER. E.g.:
        MillicomInternational
        18-feb-2026 19-feb-2026 6595348433 Aandelen USD Verkoop SLUITEN ...
        CellularSA
    """
    parts = []

    # Check previous line for instrument name fragment
    if current_idx > 0:
        prev_line = lines[current_idx - 1].strip()
        # Instrument fragments are text-only lines (no dates, no numbers-heavy content)
        if (prev_line and
            not re.match(r'\d{1,2}-[a-z]{3}-\d{4}', prev_line, re.IGNORECASE) and
            not re.match(r'Transacti', prev_line, re.IGNORECASE) and
            not re.match(r'Totaal', prev_line, re.IGNORECASE) and
            not re.match(r'Pagina', prev_line, re.IGNORECASE) and
            not re.match(r'Saxo', prev_line, re.IGNORECASE) and
            not re.match(r'Verslagperiode', prev_line, re.IGNORECASE) and
            not re.match(r'Inleiding', prev_line, re.IGNORECASE) and
            re.search(r'[A-Za-z]{2,}', prev_line)):
            parts.append(prev_line)

    # Check next line for instrument name fragment
    if current_idx < len(lines) - 1:
        next_line = lines[current_idx + 1].strip()
        if (next_line and
            not re.match(r'\d{1,2}-[a-z]{3}-\d{4}', next_line, re.IGNORECASE) and
            not re.match(r'Transacti', next_line, re.IGNORECASE) and
            not re.match(r'Totaal', next_line, re.IGNORECASE) and
            not re.match(r'Pagina', next_line, re.IGNORECASE) and
            not re.match(r'Saxo', next_line, re.IGNORECASE) and
            not re.match(r'Verslagperiode', next_line, re.IGNORECASE) and
            not re.match(r'Inleiding', next_line, re.IGNORECASE) and
            re.search(r'[A-Za-z]{2,}', next_line)):
            parts.append(next_line)

    return ' '.join(parts) if parts else 'Unknown'


def extract_saxo_transactions(text: str) -> ExtractionResult:
    """
    Extract transactions from Saxo Bank statement (Dutch/Flemish format).
    Auto-detects between two Saxo report formats:

    1. Transactierapport (preferred): Block-based format with "Transactie <Name> Koop/Verkoop..."
    2. TransactionBalance (legacy): Columnar format with "Transacties - Zelf Beleggen" sections

    CRITICAL: TOB is calculated on shares x price (gross trade value, BEFORE costs).
    The Boekingsbedrag includes broker costs and must NOT be used as the tax base.
    """
    # Try the Transactierapport format first (preferred, no line-splitting issues)
    if 'Transactierapport' in text or re.search(r'Transactie\s+\S+.*?(Koop|Verkoop)', text):
        result = _extract_saxo_transactierapport(text)
        if result.success:
            return result

    # Fall back to TransactionBalance columnar format
    result = _extract_saxo_transaction_balance(text)
    if result.success:
        return result

    # Last resort: alternative line-by-line extraction
    transactions, warnings = _extract_saxo_alternative(text)
    if transactions:
        return ExtractionResult(
            transactions=transactions,
            broker=Broker.SAXO_BANK,
            warnings=warnings,
            raw_text=text,
            success=True
        )

    return ExtractionResult(
        transactions=[],
        broker=Broker.SAXO_BANK,
        warnings=[],
        raw_text=text,
        success=False,
        error_message="No transactions found in Saxo Bank statement. "
                     "Check that the PDF contains stock transactions."
    )


# Dutch month abbreviations (shared across Saxo parsers)
DUTCH_MONTHS = {
    'jan': '01', 'feb': '02', 'mrt': '03', 'apr': '04',
    'mei': '05', 'jun': '06', 'jul': '07', 'aug': '08',
    'sep': '09', 'okt': '10', 'nov': '11', 'dec': '12'
}

# Supported currencies
SUPPORTED_CURRENCIES = 'EUR|USD|GBP|CAD|CHF|SEK|NOK|DKK|JPY|AUD|HKD'


def _parse_belgian_number(s: str) -> float:
    """Parse a Belgian-formatted number: 1.234,56 -> 1234.56"""
    return float(s.replace('.', '').replace(',', '.'))


def _parse_dutch_date(day: str, month_nl: str, year: str) -> str:
    """Parse a Dutch date (dd-mmm-yyyy) to YYYY-MM-DD format."""
    month_num = DUTCH_MONTHS.get(month_nl.lower(), '01')
    return f"{year}-{month_num}-{day.zfill(2)}"


def _extract_saxo_transactierapport(text: str) -> ExtractionResult:
    """
    Extract from Saxo "Transactierapport" format.

    Each transaction is a multi-line block:
        Transactie <InstrumentName> Koop<Shares>@<Price><Currency> <CashEUR> -
        Aandeelbedrag  Commissie  Boekingsbedrag  Totalekosten
        <EUR>          <EUR>      <OrigCurrency>  <EUR>
        Omrekeningskoers  Transactie-ID  ISIN
        <rate>            <id>           <isin>

    CRITICAL: We use shares x price as the amount (gross, before costs).
    The currency is the instrument/trading currency, NOT the account currency.
    """
    transactions = []
    warnings = []
    lines = text.split('\n')

    current_date = None

    # Date line pattern: "25-feb-2026 3,77 49,76" or "05-jan-2026 -222,17 35,67"
    date_line_pattern = re.compile(
        r'^(\d{1,2})-([a-z]{3})-(\d{4})\s+',
        re.IGNORECASE
    )

    # Transaction line: "Transactie <Name> <Koop|Verkoop><Shares>@<Price><Currency> <CashAmount> -"
    # Examples:
    #   Transactie RayonierAdvancedMaterialsInc. Koop14200@9,40USD -113.078,34 -
    #   Transactie DolePLC Verkoop-9150@14,60USD 113.082,11 -
    #   Transactie MillicomInternationalCellularSA Verkoop-3000@56,68USD 144.991,04 -
    transaction_pattern = re.compile(
        r'Transactie\s+'
        r'(.+?)\s+'                                               # Instrument name
        r'(Koop|Verkoop)'                                         # Buy/Sell
        r'(-?[\d.]+)'                                             # Shares (may have . as thousands sep)
        r'@'
        r'([\d,.]+)'                                              # Price (Belgian format)
        r'(' + SUPPORTED_CURRENCIES + r')'                        # Currency
        r'\s',
        re.IGNORECASE
    )

    for line in lines:
        line = line.strip()

        # Skip non-trade lines (dividends, corporate actions, interest, etc.)
        if any(skip in line for skip in ['Corporateaction', 'Cashbedrag', 'Debetrente',
                                          'creditrente', 'bewaarrente']):
            continue

        # Track current date from date header lines
        date_match = date_line_pattern.match(line)
        if date_match:
            day, month_nl, year = date_match.groups()
            current_date = _parse_dutch_date(day, month_nl, year)
            continue

        # Try to match a transaction line
        trans_match = transaction_pattern.search(line)
        if trans_match and current_date:
            instrument, trans_type_nl, shares_str, price_str, currency = trans_match.groups()

            shares = abs(int(shares_str.replace('.', '')))
            price = _parse_belgian_number(price_str)

            # CRITICAL: amount = shares x price (gross, BEFORE costs)
            amount = round(shares * price, 2)

            trans_type = 'Sell' if trans_type_nl.lower() == 'verkoop' else 'Buy'

            transactions.append(Transaction(
                date=current_date,
                broker=Broker.SAXO_BANK.value,
                stock=instrument.strip(),
                trans_type=trans_type,
                shares=shares,
                currency=currency.upper(),
                amount=amount,
                original_line=line
            ))

    return ExtractionResult(
        transactions=transactions,
        broker=Broker.SAXO_BANK,
        warnings=warnings,
        raw_text=text,
        success=len(transactions) > 0
    )


def _extract_saxo_transaction_balance(text: str) -> ExtractionResult:
    """
    Extract from Saxo "Transactie- en saldorapport" (TransactionBalance) format.

    Columnar format with Transacties sections per account.
    Columns: Date, ValDate, TransID, Product, Instrument, Currency, Type,
             Action, Aantal, Koers, ConvRate, P/L, Boekingsbedrag, Costs

    CRITICAL: We use shares x price (Aantal x Koers) as the amount, NOT the
    Boekingsbedrag which includes broker costs. Currency = instrument currency.
    """
    transactions = []
    warnings = []

    lines = text.split('\n')

    in_transactions_section = False

    # Pattern to detect account sections
    account_section_pattern = re.compile(
        r'Transacties.*?\(\d+/\d+([A-Z]{3})\).*?,\s*([A-Z]{3})',
        re.IGNORECASE
    )

    # Transaction with instrument on SAME line
    saxo_trans_pattern = re.compile(
        r'(\d{1,2})-([a-z]{3})-(\d{4})\s+'                           # Transaction date
        r'(\d{1,2})-([a-z]{3})-(\d{4})\s+'                           # Value date
        r'(\d+)\s+'                                                   # Transaction ID
        r'(Aandelen|Effecten)\s+'                                     # Product type
        r'(.+?)\s+'                                                   # Instrument name
        r'(' + SUPPORTED_CURRENCIES + r')\s+'                         # Instrument currency
        r'(Verkoop|Koop)\s+'                                          # Type
        r'(SLUITEN|OPENING)\s+'                                       # Open/Close
        r'(-?[\d.]+)\s+'                                              # Aantal (shares)
        r'([\d,.]+)\s+'                                               # Koers (price)
    , re.IGNORECASE)

    # Transaction with instrument SPLIT across lines (long names)
    saxo_trans_no_inst_pattern = re.compile(
        r'(\d{1,2})-([a-z]{3})-(\d{4})\s+'                           # Transaction date
        r'(\d{1,2})-([a-z]{3})-(\d{4})\s+'                           # Value date
        r'(\d+)\s+'                                                   # Transaction ID
        r'(Aandelen|Effecten)\s+'                                     # Product type
        r'(' + SUPPORTED_CURRENCIES + r')\s+'                         # Currency (directly after product)
        r'(Verkoop|Koop)\s+'                                          # Type
        r'(SLUITEN|OPENING)\s+'                                       # Open/Close
        r'(-?[\d.]+)\s+'                                              # Aantal (shares)
        r'([\d,.]+)\s+'                                               # Koers (price)
    , re.IGNORECASE)

    for i, line in enumerate(lines):
        line = line.strip()

        # Check for account section header
        if account_section_pattern.search(line):
            in_transactions_section = True
            continue

        # Alternative section detection
        if 'Transacties' in line and 'Beleggen' in line:
            in_transactions_section = True
            continue

        if not in_transactions_section:
            continue

        if 'Cashbedrag' in line or 'Totaal' in line:
            continue

        if any(skip in line for skip in ['Cashdividend', 'aandelensplitsing', 'bewaarloon',
                                          'Debetrente', 'creditrente', 'bewaarrente']):
            continue

        # Try pattern with instrument on the same line
        match = saxo_trans_pattern.search(line)
        if match:
            groups = match.groups()
            trans_day, trans_month, trans_year = groups[0], groups[1], groups[2]
            instrument = groups[8]
            inst_currency = groups[9]
            trans_type_nl = groups[10]
            aantal_str = groups[12]
            koers_str = groups[13]

            date = _parse_dutch_date(trans_day, trans_month, trans_year)
            shares = abs(int(aantal_str.replace('.', '').replace(',', '.')))
            price = _parse_belgian_number(koers_str)
            amount = round(shares * price, 2)
            trans_type = 'Sell' if trans_type_nl.lower() == 'verkoop' else 'Buy'

            transactions.append(Transaction(
                date=date,
                broker=Broker.SAXO_BANK.value,
                stock=instrument.strip(),
                trans_type=trans_type,
                shares=shares,
                currency=inst_currency.upper(),
                amount=amount,
                original_line=line
            ))
            continue

        # Try pattern with instrument split across lines
        match_no_inst = saxo_trans_no_inst_pattern.search(line)
        if match_no_inst:
            groups = match_no_inst.groups()
            trans_day, trans_month, trans_year = groups[0], groups[1], groups[2]
            inst_currency = groups[8]
            trans_type_nl = groups[9]
            aantal_str = groups[11]
            koers_str = groups[12]

            instrument = _get_instrument_from_adjacent_lines(lines, i)
            date = _parse_dutch_date(trans_day, trans_month, trans_year)
            shares = abs(int(aantal_str.replace('.', '').replace(',', '.')))
            price = _parse_belgian_number(koers_str)
            amount = round(shares * price, 2)
            trans_type = 'Sell' if trans_type_nl.lower() == 'verkoop' else 'Buy'

            transactions.append(Transaction(
                date=date,
                broker=Broker.SAXO_BANK.value,
                stock=instrument.strip(),
                trans_type=trans_type,
                shares=shares,
                currency=inst_currency.upper(),
                amount=amount,
                original_line=line
            ))

    return ExtractionResult(
        transactions=transactions,
        broker=Broker.SAXO_BANK,
        warnings=warnings,
        raw_text=text,
        success=len(transactions) > 0
    )


def _extract_saxo_alternative(text: str) -> Tuple[List[Transaction], List[str]]:
    """
    Alternative extraction for Saxo statements
    Uses line-by-line parsing with more flexible patterns
    """
    transactions = []
    warnings = []
    
    dutch_months = {
        'jan': '01', 'feb': '02', 'mrt': '03', 'apr': '04', 
        'mei': '05', 'jun': '06', 'jul': '07', 'aug': '08',
        'sep': '09', 'okt': '10', 'nov': '11', 'dec': '12'
    }
    
    lines = text.split('\n')
    current_currency = 'EUR'  # Default
    
    for i, line in enumerate(lines):
        # Track account currency from section headers
        currency_section_match = re.search(r'\),\s*(EUR|USD|GBP|CAD|CHF|SEK|NOK|DKK|JPY|AUD)\s*$', line, re.IGNORECASE)
        if currency_section_match:
            current_currency = currency_section_match.group(1).upper()
        
        # Look for transaction patterns
        # Date pattern: dd-mmm-yyyy
        date_match = re.search(r'(\d{1,2})-([a-z]{3})-(\d{4})', line, re.IGNORECASE)
        if not date_match:
            continue
        
        # Must have Aandelen (stocks) and Verkoop/Koop
        if 'Aandelen' not in line:
            continue
        if 'Verkoop' not in line and 'Koop' not in line:
            continue
        
        # Skip cash transactions
        if 'Cashbedrag' in line:
            continue
        
        # Extract date
        day, month_nl, year = date_match.groups()
        month_num = dutch_months.get(month_nl.lower(), '01')
        date = f"{year}-{month_num}-{day.zfill(2)}"
        
        # Determine type
        trans_type = 'Sell' if 'Verkoop' in line else 'Buy'
        
        # Extract instrument name - appears between Aandelen and (EUR|USD)
        inst_match = re.search(r'Aandelen\s+(.+?)\s+(EUR|USD)', line)
        if inst_match:
            instrument = inst_match.group(1).strip()
        else:
            instrument = "Unknown"
        
        # Extract amount - look for Belgian format number patterns
        # The Boekingsbedrag is typically the largest positive number
        amounts = re.findall(r'(?<![\d-])(\d{1,3}(?:\.\d{3})*,\d{2})(?!\d)', line)
        if amounts:
            # Parse all amounts and take the appropriate one
            parsed_amounts = []
            for amt in amounts:
                amt_float = float(amt.replace('.', '').replace(',', '.'))
                parsed_amounts.append(amt_float)
            
            # The Boekingsbedrag is usually positive and one of the larger values
            positive_amounts = [a for a in parsed_amounts if a > 0]
            if positive_amounts:
                amount_value = max(positive_amounts)
            else:
                amount_value = max(parsed_amounts) if parsed_amounts else 0
        else:
            continue  # Can't find amount
        
        # Extract shares - look for negative number for sells
        shares_match = re.search(r'(-?\d{1,3}(?:\.\d{3})*)\s+[\d,]+\s+[\d,]+', line)
        if shares_match:
            shares_str = shares_match.group(1).replace('.', '')
            shares = abs(int(shares_str))
        else:
            shares = 0
        
        if amount_value > 0 and shares > 0:
            transactions.append(Transaction(
                date=date,
                broker=Broker.SAXO_BANK.value,
                stock=instrument,
                trans_type=trans_type,
                shares=shares,
                currency=current_currency,
                amount=amount_value,
                original_line=line
            ))
    
    return transactions, warnings


# =============================================================================
# SAXO BANK EXCEL EXTRACTION
# =============================================================================

def extract_saxo_excel(file_path: str) -> ExtractionResult:
    """
    Extract transactions from Saxo Bank Excel export.

    Uses the '_Transacties' sheet which contains clean, structured trade data:
    - Traded Quantity (signed: positive=buy, negative=sell)
    - Prijs (price at full precision)
    - Verhandelde waarde (= shares x price, gross trade value BEFORE costs)
    - Instrumentvaluta (trading currency)
    - Instrument (clean name)
    - Aangepaste transactiedatum (trade date as datetime)

    This is the most reliable Saxo extraction method — no PDF parsing needed.
    """
    from openpyxl import load_workbook

    transactions = []
    warnings = []

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        return ExtractionResult(
            transactions=[],
            broker=Broker.SAXO_BANK,
            warnings=[f"Failed to open Excel file: {e}"],
            raw_text="",
            success=False,
            error_message=f"Failed to open Excel file: {e}"
        )

    # Use _Transacties sheet (trade-only data) if available, fall back to Transacties
    if '_Transacties' in wb.sheetnames:
        ws = wb['_Transacties']
        return _parse_saxo_trades_sheet(ws, warnings)
    elif 'Transacties' in wb.sheetnames:
        ws = wb['Transacties']
        return _parse_saxo_main_sheet(ws, warnings)
    else:
        return ExtractionResult(
            transactions=[],
            broker=Broker.SAXO_BANK,
            warnings=[f"No recognized sheet found. Available: {wb.sheetnames}"],
            raw_text="",
            success=False,
            error_message="Excel file does not contain a 'Transacties' or '_Transacties' sheet."
        )


def _parse_saxo_trades_sheet(ws, warnings: list) -> ExtractionResult:
    """
    Parse the _Transacties sheet (trade-specific, clean data).

    Key columns:
    - Aangepaste transactiedatum (col index found dynamically)
    - Acties (action string, to filter out stock splits)
    - Traded Quantity (signed)
    - Verhandelde waarde (gross trade value = shares x price)
    - Instrumentvaluta (currency)
    - Instrument (stock name)
    """
    transactions = []
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        return ExtractionResult(
            transactions=[], broker=Broker.SAXO_BANK, warnings=warnings,
            raw_text="", success=False,
            error_message="No data rows in _Transacties sheet."
        )

    # Build column index from header row
    headers = [str(h).strip().replace('\xa0', ' ') if h else '' for h in rows[0]]
    col = {h: i for i, h in enumerate(headers)}

    # Required columns
    required = ['Aangepaste transactiedatum', 'Acties', 'Traded Quantity',
                'Verhandelde waarde', 'Instrumentvaluta', 'Instrument']
    missing = [r for r in required if r not in col]
    if missing:
        return ExtractionResult(
            transactions=[], broker=Broker.SAXO_BANK, warnings=warnings,
            raw_text="", success=False,
            error_message=f"Missing columns in _Transacties sheet: {missing}"
        )

    for row in rows[1:]:
        acties = str(row[col['Acties']] or '')

        # Skip non-trade rows (stock splits, etc.)
        if 'aandelensplitsing' in acties.lower():
            continue

        date_val = row[col['Aangepaste transactiedatum']]
        quantity = row[col['Traded Quantity']]
        traded_value = row[col['Verhandelde waarde']]
        currency = row[col['Instrumentvaluta']]
        instrument = row[col['Instrument']]

        # Validate required fields
        if not all([date_val, quantity, traded_value, currency, instrument]):
            continue

        # Parse date
        if hasattr(date_val, 'strftime'):
            date_str = date_val.strftime('%Y-%m-%d')
        else:
            continue

        shares = abs(int(quantity))
        # Verhandelde waarde = gross trade value (shares x price), already correct
        amount = abs(float(traded_value))
        trans_type = 'Sell' if int(quantity) < 0 else 'Buy'

        transactions.append(Transaction(
            date=date_str,
            broker=Broker.SAXO_BANK.value,
            stock=str(instrument).strip(),
            trans_type=trans_type,
            shares=shares,
            currency=str(currency).upper(),
            amount=round(amount, 2),
            original_line=acties
        ))

    return ExtractionResult(
        transactions=transactions,
        broker=Broker.SAXO_BANK,
        warnings=warnings,
        raw_text="",
        success=len(transactions) > 0
    )


def _parse_saxo_main_sheet(ws, warnings: list) -> ExtractionResult:
    """
    Parse the main Transacties sheet (has all transaction types mixed together).
    Filter to only 'Transactie' type rows (excludes dividends, corporate actions).
    Uses shares x price from the Acties column as the amount.
    """
    transactions = []
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        return ExtractionResult(
            transactions=[], broker=Broker.SAXO_BANK, warnings=warnings,
            raw_text="", success=False,
            error_message="No data rows in Transacties sheet."
        )

    headers = [str(h).strip().replace('\xa0', ' ') if h else '' for h in rows[0]]
    col = {h: i for i, h in enumerate(headers)}

    required = ['Transactiedatum', 'Transactietype', 'Acties', 'Instrument',
                'Instrumentvaluta']
    missing = [r for r in required if r not in col]
    if missing:
        return ExtractionResult(
            transactions=[], broker=Broker.SAXO_BANK, warnings=warnings,
            raw_text="", success=False,
            error_message=f"Missing columns in Transacties sheet: {missing}"
        )

    # Pattern to parse action string: "Koop 14200 @ 9.40 USD" or "Verkoop -9150 @ 14.60 USD"
    action_pattern = re.compile(
        r'(Koop|Verkoop)\s+(-?\d+)\s+@\s+([\d.]+)\s+([A-Z]{3})',
        re.IGNORECASE
    )

    for row in rows[1:]:
        trans_type_raw = str(row[col['Transactietype']] or '')

        # Only process actual trades
        if trans_type_raw != 'Transactie':
            continue

        acties = str(row[col['Acties']] or '')
        action_match = action_pattern.search(acties)
        if not action_match:
            continue

        trans_type_nl, shares_str, price_str, currency = action_match.groups()

        date_val = row[col['Transactiedatum']]
        instrument = row[col['Instrument']]

        if not all([date_val, instrument]):
            continue

        if hasattr(date_val, 'strftime'):
            date_str = date_val.strftime('%Y-%m-%d')
        else:
            continue

        shares = abs(int(shares_str))
        price = float(price_str)
        amount = round(shares * price, 2)
        trans_type = 'Sell' if trans_type_nl.lower() == 'verkoop' else 'Buy'

        transactions.append(Transaction(
            date=date_str,
            broker=Broker.SAXO_BANK.value,
            stock=str(instrument).strip(),
            trans_type=trans_type,
            shares=shares,
            currency=currency.upper(),
            amount=amount,
            original_line=acties
        ))

    return ExtractionResult(
        transactions=transactions,
        broker=Broker.SAXO_BANK,
        warnings=warnings,
        raw_text="",
        success=len(transactions) > 0
    )


# =============================================================================
# ECB RATE FETCHING
# =============================================================================

def fetch_ecb_rates(dates_needed: set) -> Dict[str, Dict[str, float]]:
    """
    Fetch ECB exchange rates from XML feed
    
    Args:
        dates_needed: Set of date strings in 'YYYY-MM-DD' format
    
    Returns:
        dict: {date: {currency: rate}}
        
    Raises:
        ExtractionError: If rates cannot be fetched
    """
    url = "https://www.ecb.europa.eu/stats/eurofxref/eurofxref-hist.xml"
    
    try:
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        xml_content = response.content
    except requests.exceptions.Timeout:
        raise ExtractionError(
            "ECB rate server timeout. Please try again.",
            "ecb_timeout"
        )
    except requests.exceptions.ConnectionError:
        raise ExtractionError(
            "Cannot connect to ECB rate server. Check internet connection.",
            "ecb_connection_error"
        )
    except Exception as e:
        raise ExtractionError(
            f"Failed to fetch ECB rates: {str(e)}",
            "ecb_fetch_error",
            {"original_error": str(e)}
        )
    
    return parse_ecb_xml(xml_content, dates_needed)


def parse_ecb_xml(xml_content: bytes, dates_needed: set) -> Dict[str, Dict[str, float]]:
    """Parse ECB XML and extract rates for needed dates"""
    ns = {
        'gesmes': 'http://www.gesmes.org/xml/2002-08-01',
        'default': 'http://www.ecb.int/vocabulary/2002-08-01/eurofxref'
    }
    
    try:
        root = ET.fromstring(xml_content)
    except ET.ParseError as e:
        raise ExtractionError(
            f"Invalid ECB XML format: {str(e)}",
            "ecb_parse_error"
        )
    
    rates = {}
    all_available_dates = []
    
    # First pass: collect all available dates and their rates
    for date_cube in root.findall('.//default:Cube[@time]', ns):
        date = date_cube.get('time')
        all_available_dates.append(date)
        
        if date in dates_needed:
            rates[date] = {'EUR': 1.0}  # EUR to EUR is always 1.0
            for curr_cube in date_cube.findall('default:Cube[@currency]', ns):
                currency = curr_cube.get('currency')
                rate = float(curr_cube.get('rate'))
                rates[date][currency] = rate
    
    # Build a lookup of all available dates and their rates for efficient fallback
    all_rates_by_date = {}
    for date_cube in root.findall('.//default:Cube[@time]', ns):
        d = date_cube.get('time')
        day_rates = {'EUR': 1.0}
        for curr_cube in date_cube.findall('default:Cube[@currency]', ns):
            currency = curr_cube.get('currency')
            rate = float(curr_cube.get('rate'))
            day_rates[currency] = rate
        all_rates_by_date[d] = day_rates

    # Handle missing dates (weekends/holidays) - use previous business day
    for date in dates_needed:
        if date not in rates:
            rates[date] = _get_rate_with_fallback(date, all_rates_by_date)

    return rates


def _get_rate_with_fallback(date: str, all_rates: Dict[str, Dict[str, float]]) -> Dict[str, float]:
    """Get rate for date, falling back to previous business day if needed"""
    current_date = datetime.strptime(date, '%Y-%m-%d')

    # Try up to 7 days back (to handle long holiday periods)
    for i in range(1, 8):
        prev_date = (current_date - timedelta(days=i)).strftime('%Y-%m-%d')
        if prev_date in all_rates:
            return all_rates[prev_date]

    raise ExtractionError(
        f"No ECB rate found for {date} or 7 days prior. "
        "This date may be outside the ECB historical data range.",
        "ecb_rate_missing",
        {"date": date}
    )


# =============================================================================
# TRANSACTION GROUPING
# =============================================================================

def group_transactions(transactions: List[Transaction]) -> List[Dict[str, Any]]:
    """
    Group transactions by date + stock + type
    
    Rules:
    - Same-side transactions (multiple buys OR multiple sells on same day) = GROUP into 1 transaction
    - Opposite-side transactions (day trades with both buy AND sell) = KEEP SEPARATE
    - Belgian law taxes both sides of a day trade separately
    
    Args:
        transactions: List of Transaction objects
        
    Returns:
        List of grouped transaction dictionaries
    """
    grouped = defaultdict(lambda: {
        'shares': 0,
        'amount': 0.0,
        'transactions': []
    })
    
    for t in transactions:
        # Group key: date + broker + stock + type + currency
        key = (t.date, t.broker, t.stock, t.trans_type, t.currency)
        grouped[key]['shares'] += t.shares
        grouped[key]['amount'] += t.amount
        grouped[key]['transactions'].append(t)
    
    result = []
    for key, data in grouped.items():
        date, broker, stock, trans_type, currency = key
        result.append({
            'date': date,
            'broker': broker,
            'stock': stock,
            'type': trans_type,
            'shares': data['shares'],
            'currency': currency,
            'amount': round(data['amount'], 2),
            'grouped_count': len(data['transactions'])
        })
    
    # Sort by date, then broker, then stock
    return sorted(result, key=lambda x: (x['date'], x['broker'], x['stock']))


# =============================================================================
# TOB CALCULATION
# =============================================================================

def calculate_tob(transactions: List[Dict], ecb_rates: Dict[str, Dict[str, float]]) -> List[Dict]:
    """
    Calculate TOB for all transactions
    
    TOB Rate: 0.35% (0.0035)
    Conversion: EUR = Amount / ECB_Rate
    
    Args:
        transactions: List of grouped transaction dicts
        ecb_rates: ECB rate dictionary
        
    Returns:
        List of transactions with calculated TOB
    """
    results = []
    
    for t in transactions:
        date = t['date']
        currency = t['currency']
        amount = t['amount']
        
        # Get ECB rate
        if date not in ecb_rates:
            raise ExtractionError(
                f"No ECB rate data for date {date}",
                "missing_rate",
                {"date": date}
            )
        
        if currency not in ecb_rates[date]:
            raise ExtractionError(
                f"No ECB rate for {currency} on {date}",
                "missing_currency_rate",
                {"currency": currency, "date": date}
            )
        
        rate = ecb_rates[date][currency]
        
        # Convert to EUR
        # ECB rate format: 1 EUR = rate units of foreign currency
        # So: EUR = Foreign / Rate
        # For EUR: rate is 1.0, so eur_amount = amount
        eur_amount = round(amount / rate, 2)
        
        # Calculate TOB at 0.35%
        tob = round(eur_amount * 0.0035, 2)
        
        # Apply maximum cap (EUR 1,600 per transaction)
        tob = min(tob, 1600.00)
        
        results.append({
            **t,
            'rate': rate,
            'eur_amount': eur_amount,
            'tob': tob
        })
    
    return results


# =============================================================================
# MAIN PROCESSING FUNCTION
# =============================================================================

def process_statements(pdf_paths: List[str]) -> Dict[str, Any]:
    """
    Main processing function
    
    Args:
        pdf_paths: List of paths to PDF files
        
    Returns:
        Dictionary with:
        - transactions: List of processed transactions with TOB
        - total_eur: Total EUR amount
        - total_tob: Total TOB tax
        - ecb_rates: ECB rates used
        - warnings: Any warnings generated
        - brokers: List of brokers found
        
    Raises:
        ExtractionError: If processing fails
    """
    all_transactions = []
    all_warnings = []
    brokers_found = set()
    
    if not pdf_paths:
        raise ExtractionError(
            "No files provided",
            "no_files"
        )

    # Extract transactions from all files (PDFs and Excel)
    for file_path in pdf_paths:
        is_excel = file_path.lower().endswith('.xlsx')

        if is_excel:
            # Saxo Excel export — structured data, most reliable
            result = extract_saxo_excel(file_path)
            brokers_found.add(Broker.SAXO_BANK.value)
        else:
            # PDF file — extract text and detect broker
            try:
                text = extract_text_from_pdf(file_path)
            except ExtractionError:
                raise
            except Exception as e:
                raise ExtractionError(
                    f"Failed to read PDF file: {file_path}",
                    "pdf_read_error",
                    {"path": file_path, "error": str(e)}
                )

            broker = detect_broker(text)
            brokers_found.add(broker.value)

            if broker == Broker.INTERACTIVE_BROKERS:
                result = extract_ib_transactions(text)
            elif broker == Broker.SAXO_BANK:
                result = extract_saxo_transactions(text)
            else:
                raise ExtractionError(
                    f"Unknown broker format in {file_path}. "
                    "Supported brokers: Interactive Brokers, Saxo Bank",
                    "unknown_broker",
                    {"path": file_path}
                )
        
        if not result.success:
            raise ExtractionError(
                result.error_message,
                "extraction_failed",
                {"path": pdf_path, "broker": broker.value}
            )
        
        all_transactions.extend(result.transactions)
        all_warnings.extend(result.warnings)
    
    if not all_transactions:
        raise ExtractionError(
            "No transactions found in any of the uploaded PDFs. "
            "Please verify the PDFs contain stock transactions.",
            "no_transactions"
        )
    
    # Group transactions
    grouped = group_transactions(all_transactions)
    
    # Get unique dates for ECB rate fetching
    dates_needed = set(t['date'] for t in grouped)
    
    # Fetch ECB rates
    ecb_rates = fetch_ecb_rates(dates_needed)
    
    # Calculate TOB
    results = calculate_tob(grouped, ecb_rates)
    
    # Calculate totals
    total_eur = sum(r['eur_amount'] for r in results)
    total_tob = sum(r['tob'] for r in results)
    
    return {
        'transactions': results,
        'total_eur': round(total_eur, 2),
        'total_tob': round(total_tob, 2),
        'ecb_rates': ecb_rates,
        'warnings': all_warnings,
        'brokers': list(brokers_found),
        'transaction_count': len(results)
    }


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def format_belgian_number(num: float, decimals: int = 2) -> str:
    """
    Format number in Belgian style
    Belgian format: 1.234,56 (period for thousands, comma for decimals)

    Args:
        num: Number to format
        decimals: Number of decimal places

    Returns:
        Belgian-formatted string
    """
    sign = "-" if num < 0 else ""
    num = abs(num)

    if decimals == 0:
        formatted = f"{int(round(num)):,}".replace(',', '.')
    else:
        int_part = int(num)
        dec_part = round(num - int_part, decimals)

        # Format integer with periods
        int_str = f"{int_part:,}".replace(',', '.')

        # Format decimal with comma
        dec_str = f"{dec_part:.{decimals}f}"[1:]  # Get ".XX" part
        dec_str = dec_str.replace('.', ',')  # Change to ",XX"

        formatted = int_str + dec_str

    return sign + formatted


def validate_transaction_data(transactions: List[Dict]) -> List[str]:
    """
    Validate transaction data for common issues
    
    Args:
        transactions: List of transaction dictionaries
        
    Returns:
        List of validation warnings
    """
    warnings = []
    
    for i, t in enumerate(transactions):
        # Check for zero amounts
        if t.get('amount', 0) <= 0:
            warnings.append(f"Transaction {i+1}: Zero or negative amount for {t.get('stock', 'unknown')}")
        
        # Check for zero shares
        if t.get('shares', 0) <= 0:
            warnings.append(f"Transaction {i+1}: Zero or negative shares for {t.get('stock', 'unknown')}")
        
        # Check for valid date
        try:
            datetime.strptime(t.get('date', ''), '%Y-%m-%d')
        except ValueError:
            warnings.append(f"Transaction {i+1}: Invalid date format for {t.get('stock', 'unknown')}")
        
        # Check for known currency
        known_currencies = {'EUR', 'USD', 'GBP', 'CAD', 'AUD', 'JPY', 'CHF', 'SEK', 'NOK', 'DKK', 'HKD'}
        if t.get('currency', '') not in known_currencies:
            warnings.append(f"Transaction {i+1}: Unknown currency {t.get('currency', '')} for {t.get('stock', 'unknown')}")
    
    return warnings
